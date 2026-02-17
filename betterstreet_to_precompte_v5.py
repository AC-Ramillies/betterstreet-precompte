#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BetterStreet -> Précompte (Planning ouvriers)

Objectif:
- Lire un export BetterStreet CSV (séparateur ;) potentiellement cassé (retours lignes / ; non quotés).
- Reconstruire 1 intervention = 1 ligne, détectée via ID commençant par BE-.
- Extraire proprement les champs (priorité au mapping par en-tête quand la ligne est "alignée").
- Filtrer par année (planif si dispo, sinon "Créé le").
- Générer un Excel:
  - Feuille "Planning" (tri chronologique)
  - Feuille "Anomalies"
  - Surlignage:
      * Rouge: parsing/missing (start/end) + Agents/Équipes missing
      * Bleu: Heure début < 08:00
- Journaliser un contrôle qualité simple en console.

Usage:
  python betterstreet_to_precompte_v5.py Export_Planning_BetterStreet.csv 2025 [sortie.xlsx]
"""

from __future__ import annotations

import sys
import re
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime, timedelta, date, time
from typing import Optional, Dict, List, Tuple, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------
# Regex / heuristiques
# ---------------------------
RE_DT = re.compile(r"^\d{2}[-/]\d{2}[-/]\d{2,4}\s+\d{2}:\d{2}(:\d{2})?$")
RE_DATE = re.compile(r"^\d{2}[-/]\d{2}[-/]\d{2,4}$")

RE_BS_ID = re.compile(r"^BE-\d{3,5}-", re.IGNORECASE)
RE_STREET = re.compile(
    r"\b(rue|avenue|av\.?|chaussée|ch\.?|place|impasse|allée|chemin|route|sentier|clos|quai|boulevard|bd\.?)\b",
    re.IGNORECASE,
)


DT_FORMATS = [
    "%d-%m-%y %H:%M",
    "%d-%m-%Y %H:%M",
    "%d/%m/%y %H:%M",
    "%d/%m/%Y %H:%M",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%d-%m-%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
]
DATE_ONLY_FORMATS = [
    "%d-%m-%y",
    "%d-%m-%Y",
    "%d/%m/%y",
    "%d/%m/%Y",
]


def is_empty(v: Any) -> bool:
    return v is None or str(v).strip() == ""


def looks_like_betterstreet_id(v: Any) -> bool:
    if v is None:
        return False
    t = str(v).strip()
    return t.startswith("BE-") or bool(RE_BS_ID.match(t))


def looks_like_address(v: Any) -> bool:
    """
    Détection stricte (anti-faux positifs):
    - refuse explicitement les IDs BetterStreet
    - exige un mot de rue (rue/avenue/chaussée/...) + un chiffre
    """
    if v is None:
        return False
    t = str(v).strip()
    if len(t) < 10:
        return False
    if looks_like_betterstreet_id(t):
        return False
    if not any(c.isalpha() for c in t):
        return False
    if not RE_STREET.search(t):
        return False
    if not any(c.isdigit() for c in t):
        return False
    return True


def parse_dt(val: str) -> datetime:
    s = val.strip()
    for fmt in DT_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise ValueError(f"Date/heure non reconnue: {s}")


def parse_date_only(val: str) -> Optional[date]:
    s = val.strip()
    for fmt in DATE_ONLY_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def in_year(dt: Optional[datetime], year: int) -> bool:
    return bool(dt) and dt.year == year


# ---------------------------
# Lecture / reconstruction lignes cassées
# ---------------------------
def safe_read_text(csv_path: Path, encodings=("utf-8-sig", "utf-8", "cp1252")) -> Tuple[str, str]:
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            return csv_path.read_text(encoding=enc, errors="strict"), enc
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Impossible de lire le fichier ({csv_path}) avec encodages {encodings}: {last_err}")


def rebuild_records_by_id(lines: List[str], delimiter: str = ";") -> List[str]:
    """
    Recolle les retours à la ligne dans les champs textes.
    Nouveau record si le 1er champ commence par BE-.
    """
    if not lines:
        return []

    i = 0
    while i < len(lines) and not lines[i].strip():
        i += 1
    if i >= len(lines):
        return []

    header = lines[i].rstrip("\n")
    rebuilt = [header]

    current: Optional[str] = None
    for raw in lines[i + 1 :]:
        line = raw.rstrip("\n")
        if not line.strip():
            continue

        first = line.split(delimiter, 1)[0]
        first = first.replace("\ufeff", "").strip().strip('"').strip("'")

        if first.startswith("BE-"):
            if current is not None:
                rebuilt.append(current)
            current = line
        else:
            if current is None:
                continue
            current += " " + line

    if current is not None:
        rebuilt.append(current)

    return rebuilt


# ---------------------------
# Extraction champs
# ---------------------------
@dataclass
class ExtractResult:
    fields: Dict[str, Optional[str]]
    aligned: bool  # mapping par en-tête (fiable) ?


def extract_fields(record_line: str, header_cols: List[str], delimiter: str = ";") -> Optional[ExtractResult]:
    toks = [t.strip() for t in record_line.split(delimiter)]
    if not toks:
        return None

    rid = toks[0].replace("\ufeff", "").strip().strip('"').strip("'")
    if not rid.startswith("BE-"):
        return None

    # 1) Cas "aligné": même nb de colonnes que l'en-tête -> mapping direct (NE PAS valider/écraser)
    if len(toks) == len(header_cols):
        data = dict(zip(header_cols, toks))
        # normaliser vides
        norm = {k: (v if v != "" else None) for k, v in data.items()}
        return ExtractResult(fields=norm, aligned=True)

    # 2) Cas "cassé": fallback par ancres
    dt_idxs = [i for i, t in enumerate(toks) if t and RE_DT.match(t)]
    start_idx, end_idx = (dt_idxs[-2], dt_idxs[-1]) if len(dt_idxs) >= 2 else (None, None)

    date_idxs = [i for i, t in enumerate(toks) if t and RE_DATE.match(t)]
    created_idx = echeance_idx = None
    if start_idx is not None:
        prev = [i for i in date_idxs if i < start_idx]
        if len(prev) >= 2:
            created_idx, echeance_idx = prev[-2], prev[-1]
    else:
        if len(date_idxs) >= 2:
            created_idx, echeance_idx = date_idxs[0], date_idxs[1]

    f: Dict[str, Optional[str]] = {k: None for k in header_cols}
    f["ID"] = rid

    if created_idx is not None:
        f["Créé le"] = toks[created_idx] or None
        # Catégorie potentielle juste avant "Créé le"
        if created_idx - 1 >= 1:
            cand = toks[created_idx - 1].strip()
            if cand and not RE_DATE.match(cand) and not RE_DT.match(cand):
                f["Catégorie"] = cand
        # Description = entre ID et created_idx-1
        if created_idx > 2:
            desc = ";".join(toks[1 : created_idx - 1]).strip()
            f["Description"] = desc or None
        elif len(toks) > 1:
            f["Description"] = toks[1].strip() or None
    elif len(toks) > 1:
        f["Description"] = toks[1].strip() or None

    if echeance_idx is not None:
        f["Échéance"] = toks[echeance_idx] or None

    if start_idx is not None:
        f["Début planification"] = toks[start_idx] or None
    if end_idx is not None:
        f["Fin planification"] = toks[end_idx] or None

    # Agents/Équipes: juste après fin planif, sinon après Échéance (+3) en mode non planifié
    agents_idx = None
    if end_idx is not None and end_idx + 1 < len(toks):
        agents_idx = end_idx + 1
    elif echeance_idx is not None and echeance_idx + 3 < len(toks):
        agents_idx = echeance_idx + 3
    if agents_idx is not None:
        f["Agents/Équipes"] = toks[agents_idx] or None

    # Adresse/Bâtiment autour de l'échéance si valide, sinon fallback scan "adresse-like"
    if echeance_idx is not None:
        addr_i = echeance_idx - 1
        bat_i = echeance_idx - 2
        if 0 <= addr_i < len(toks):
            cand_addr = toks[addr_i].strip()
            if cand_addr and looks_like_address(cand_addr):
                f["Adresse"] = cand_addr
        if 0 <= bat_i < len(toks):
            cand_bat = toks[bat_i].strip()
            if cand_bat and not looks_like_betterstreet_id(cand_bat) and not looks_like_address(cand_bat) and not RE_DATE.match(cand_bat) and not RE_DT.match(cand_bat):
                f["Bâtiment/Équipement"] = cand_bat

    if not f.get("Adresse"):
        for tok in toks:
            if looks_like_address(tok):
                f["Adresse"] = tok.strip()
                break

    # Bâtiment fallback: chercher juste avant l'adresse dans une petite fenêtre
    if not f.get("Bâtiment/Équipement") and f.get("Adresse"):
        try:
            addr_pos = next(i for i, t in enumerate(toks) if t.strip() == str(f["Adresse"]).strip())
            for j in range(addr_pos - 1, max(-1, addr_pos - 8), -1):
                cand = toks[j].strip()
                if not cand:
                    continue
                if looks_like_betterstreet_id(cand) or looks_like_address(cand) or RE_DATE.match(cand) or RE_DT.match(cand):
                    continue
                if f.get("Description") and cand == f["Description"]:
                    continue
                f["Bâtiment/Équipement"] = cand
                break
        except StopIteration:
            pass

    # Consigne: prendre le dernier champ textuel "significatif" après agents_idx
    if agents_idx is not None and agents_idx + 1 < len(toks):
        tail = toks[agents_idx + 1 :]
        for tok in reversed(tail):
            t = tok.strip()
            if not t:
                continue
            if looks_like_betterstreet_id(t) or looks_like_address(t) or RE_DATE.match(t) or RE_DT.match(t):
                continue
            if f.get("Description") and t == f["Description"]:
                continue
            f["Consigne"] = t
            break

    return ExtractResult(fields=f, aligned=False)


# ---------------------------
# Excel styling
# ---------------------------
BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # < 08:00
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # parsing/missing/agents


DEFAULT_OUTPUT_NAME = "Planning_Ouvriers_Reformatte.xlsx"


def main(csv_path_str: str, year: int, out_path_str: Optional[str] = None) -> None:
    csv_path = Path(csv_path_str)
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV introuvable: {csv_path}")

    out_path = Path(out_path_str) if out_path_str else csv_path.with_name(DEFAULT_OUTPUT_NAME)

    raw_text, encoding_used = safe_read_text(csv_path)
    raw_lines = raw_text.splitlines(True)
    rebuilt_lines = rebuild_records_by_id(raw_lines)

    if not rebuilt_lines:
        raise ValueError("Aucune ligne reconstruite (CSV vide ?).")

    header_cols = [c.strip() for c in rebuilt_lines[0].split(";")]
    expected_cols = len(header_cols)

    print("Lecture OK (réparation CSV) :")
    print(f"  - encoding_used: {encoding_used}")
    print(f"  - raw_line_count: {len(raw_lines)}")
    print(f"  - rebuilt_record_count: {max(0, len(rebuilt_lines)-1)}")
    print(f"  - expected_columns: {expected_cols}")

    records: List[Dict[str, Any]] = []
    anomalies: List[Dict[str, Any]] = []

    kept = 0
    skipped_other_year = 0
    aligned_count = 0
    broken_count = 0

    for record_line in rebuilt_lines[1:]:
        ex = extract_fields(record_line, header_cols)
        if ex is None:
            continue
        f = ex.fields
        if ex.aligned:
            aligned_count += 1
        else:
            broken_count += 1

        start_raw = f.get("Début planification")
        end_raw = f.get("Fin planification")
        created_raw = f.get("Créé le")
        agents_equipes = f.get("Agents/Équipes")

        # parsing
        start_dt: Optional[datetime] = None
        end_dt: Optional[datetime] = None
        created_date: Optional[date] = None
        notes: List[str] = []

        try:
            if not is_empty(start_raw):
                start_dt = parse_dt(str(start_raw))
        except Exception as e:
            notes.append(f"Start parse error: {e}")

        try:
            if not is_empty(end_raw):
                end_dt = parse_dt(str(end_raw))
        except Exception as e:
            notes.append(f"End parse error: {e}")

        if not is_empty(created_raw):
            created_date = parse_date_only(str(created_raw))

        if start_dt is None:
            notes.append("Start missing")
        if end_dt is None:
            notes.append("End missing")
        if is_empty(agents_equipes):
            notes.append("Agents/Équipes missing")

        # Filtre année
        planned_in_year = in_year(start_dt, year) or in_year(end_dt, year)
        created_in_year = (created_date is not None and created_date.year == year)
        has_planning = (start_dt is not None) or (end_dt is not None)
        keep = planned_in_year if has_planning else created_in_year

        if not keep:
            skipped_other_year += 1
            continue

        kept += 1

        # Correction 12h forcée si fin < début
        end_dt_corrected = end_dt
        if start_dt and end_dt and end_dt < start_dt:
            end_dt_corrected = end_dt + timedelta(hours=12)
            notes.append("End time shifted +12h (12h ambiguity forced)")

        duration_td: Optional[timedelta] = None
        duration_str: Optional[str] = None
        if start_dt and end_dt_corrected:
            duration_td = end_dt_corrected - start_dt
            duration_str = str(duration_td)
            if duration_td.total_seconds() < 0:
                notes.append("Duration still negative after correction")

        notes_joined = " | ".join(notes)

        # Anomalies: inclure parsing/missing/agents + correction +12h
        if (
            "parse error" in notes_joined
            or "missing" in notes_joined
            or "End time shifted +12h" in notes_joined
            or "Duration still negative" in notes_joined
        ):
            anomalies.append({
                "ID": f.get("ID"),
                "Début planification (raw)": start_raw,
                "Fin planification (raw)": end_raw,
                "Fin planification (corrigée)": end_dt_corrected.strftime("%Y-%m-%d %H:%M") if end_dt_corrected else None,
                "Durée (corrigée)": str(duration_td) if duration_td is not None else None,
                "Notes": notes_joined,
            })

        records.append({
            "ID": f.get("ID"),
            "Description": f.get("Description"),
            "Catégorie": f.get("Catégorie"),
            "Bâtiment/Équipement": f.get("Bâtiment/Équipement"),
            "Adresse": f.get("Adresse"),
            "Date Début planification": start_dt.date() if start_dt else None,
            "Heure Début planification": start_dt.time() if start_dt else None,
            "Date Fin planification": end_dt_corrected.date() if end_dt_corrected else None,
            "Heure Fin planification": end_dt_corrected.time() if end_dt_corrected else None,
            "Consigne": f.get("Consigne"),
            "Agents/Équipes": agents_equipes,
            "Durée": duration_str,
        })

    df_out = pd.DataFrame(records)

    # TRI CHRONOLOGIQUE (clé technique)
    if not df_out.empty and "Date Début planification" in df_out.columns:
        df_out["__sort_dt"] = pd.to_datetime(
            df_out["Date Début planification"].astype(str) + " " + df_out["Heure Début planification"].astype(str),
            errors="coerce",
        )
        df_out = df_out.sort_values("__sort_dt", na_position="last").drop(columns="__sort_dt")

    df_anom = pd.DataFrame(anomalies)

    # IDs rouges = anomalies parsing/missing/agents
    red_ids = set()
    if not df_anom.empty and "Notes" in df_anom.columns and "ID" in df_anom.columns:
        notes_col = df_anom["Notes"].astype(str)
        mask = (
            notes_col.str.contains("Start parse error", na=False)
            | notes_col.str.contains("End parse error", na=False)
            | notes_col.str.contains("Start missing", na=False)
            | notes_col.str.contains("End missing", na=False)
            | notes_col.str.contains("Agents/Équipes missing", na=False)
            | notes_col.str.contains("Missing start/end", na=False)
        )
        red_ids = set(df_anom.loc[mask, "ID"].astype(str))

    # Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Planning"

    rows = list(dataframe_to_rows(df_out, index=False, header=True))
    for row_idx, excel_row in enumerate(rows, start=1):
        ws1.append(excel_row)
        if row_idx == 1:
            continue

        row_id = str(excel_row[0]) if excel_row and excel_row[0] is not None else ""

        # Rouge prioritaire
        if row_id in red_ids:
            for col in range(1, len(excel_row) + 1):
                ws1.cell(row=row_idx, column=col).fill = RED_FILL
            continue

        # Bleu si heure début < 08:00
        try:
            # Excel_row index 6 = Heure Début planification (0-based)
            heure_debut = excel_row[6]
            if heure_debut and hasattr(heure_debut, "hour") and heure_debut.hour < 8:
                for col in range(1, len(excel_row) + 1):
                    ws1.cell(row=row_idx, column=col).fill = BLUE_FILL
        except Exception:
            pass

    ws2 = wb.create_sheet("Anomalies")
    if df_anom.empty:
        ws2.append(["Aucune anomalie détectée avec les règles actuelles."])
    else:
        for r in dataframe_to_rows(df_anom, index=False, header=True):
            ws2.append(r)

    wb.save(out_path)

    # Contrôle qualité console
    def count_empty(col: str) -> int:
        if col not in df_out.columns or df_out.empty:
            return 0
        return int(df_out[col].isna().sum())

    print(f"OK: fichier généré -> {out_path}")
    print(f"INFO: Année ciblée: {year} | gardées: {kept} | ignorées (autres années): {skipped_other_year}")
    print(f"INFO: Anomalies (année {year}): {len(df_anom)}")
    print(f"INFO: Lignes alignées: {aligned_count} | Lignes cassées (fallback): {broken_count}")
    print("INFO: Champs vides (Planning):",
          f"Description={count_empty('Description')}",
          f"Catégorie={count_empty('Catégorie')}",
          f"Bât/Équip={count_empty('Bâtiment/Équipement')}",
          f"Adresse={count_empty('Adresse')}",
          f"Consigne={count_empty('Consigne')}",
          f"Agents={count_empty('Agents/Équipes')}",
          )


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python betterstreet_to_precompte_v5.py <Export_Planning_BetterStreet.csv> <annee> [sortie.xlsx]")
        sys.exit(1)

    csv_file = sys.argv[1]
    year = int(sys.argv[2])
    out_file = sys.argv[3] if len(sys.argv) >= 4 else None
    main(csv_file, year, out_file)
